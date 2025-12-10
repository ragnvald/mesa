#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
data_analysis_presentation.py - Compare analysis groups side by side.

This utility reads the GeoParquet outputs produced by data_analysis_setup.py /
data_process.py and lets the user compare two analysis groups without opening
the map interface.
"""

from __future__ import annotations

import argparse
import configparser
import datetime as dt
import locale
import math
import os
import subprocess
import sys
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import geopandas as gpd
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from openpyxl.drawing.image import Image as XLImage

try:
    import ttkbootstrap as tb
    from ttkbootstrap import ttk
    from ttkbootstrap.constants import INFO, PRIMARY, SECONDARY, SUCCESS, WARNING
except ModuleNotFoundError as exc:  # pragma: no cover - ttkbootstrap required at runtime
    raise SystemExit("ttkbootstrap is required (pip install ttkbootstrap).") from exc

try:
    locale.setlocale(locale.LC_ALL, "en_US.UTF-8")
except Exception:
    try:
        locale.setlocale(locale.LC_ALL, "")
    except Exception:
        pass

try:
    import matplotlib

    matplotlib.use("TkAgg")
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
except Exception as exc:  # pragma: no cover - matplotlib is required at runtime
    raise SystemExit("matplotlib with TkAgg support is required for this tool.") from exc

# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #

ANALYSIS_POLYGON_TABLE = "tbl_analysis_polygons.parquet"
ANALYSIS_GROUP_TABLE = "tbl_analysis_group.parquet"
ANALYSIS_FLAT_TABLE = "tbl_analysis_flat.parquet"
ANALYSIS_STACKED_TABLE = "tbl_analysis_stacked.parquet"
DEFAULT_PARQUET_SUBDIR = "output/geoparquet"
DEFAULT_ANALYSIS_GEOCODE = "basic_mosaic"
KM2_DENOMINATOR = 1_000_000.0

UNKNOWN_CODE = "UNKNOWN"
DEFAULT_SENSITIVITY_ORDER = ["A", "B", "C", "D", "E"]

DEFAULT_COLOR_FALLBACK: Dict[str, str] = {
    "A": "#005f73",
    "B": "#0a9396",
    "C": "#94d2bd",
    "D": "#ee9b00",
    "E": "#ca6702",
    UNKNOWN_CODE: "#8d99ae",
}


# --------------------------------------------------------------------------- #
# Helpers shared with the setup tool
# --------------------------------------------------------------------------- #

def apply_bootstyle(widget: Any, style: str) -> None:
    if not style:
        return
    try:
        widget.configure(bootstyle=style)
    except Exception:
        pass

def debug_log(base_dir: Path, message: str) -> None:
    """Append a timestamped message to log.txt for diagnostics."""
    try:
        ts = dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        target = base_dir / "log.txt"
        if not target.exists():
            alt = base_dir / "code" / "log.txt"
            if alt.exists():
                target = alt
        path = target.resolve()
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "a", encoding="utf-8") as fh:
            fh.write(f"[{ts}] [data_analysis_presentation] {message}\n")
    except Exception:
        pass


def resolve_base_dir(cli_path: Optional[str] = None) -> Path:
    """Mirror the directory probing logic used in the setup tool."""
    candidates: List[Path] = []
    env_base = os.environ.get("MESA_BASE_DIR")
    if env_base:
        candidates.append(Path(env_base))
    if cli_path:
        candidates.append(Path(cli_path))

    # Frozen exe check
    if getattr(sys, "frozen", False):
        exe_path = Path(sys.executable).resolve()
        candidates.extend([exe_path.parent, exe_path.parent.parent])

    here = Path(__file__).resolve()
    candidates.extend([here.parent, here.parent.parent, here.parent.parent.parent])
    cwd = Path(os.getcwd())
    candidates.extend([cwd, cwd / "code"])

    ordered: List[Path] = []
    seen: set[Path] = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except Exception:
            resolved = candidate
        if resolved not in seen:
            seen.add(resolved)
            ordered.append(resolved)

    for candidate in ordered:
        if (candidate / "config.ini").exists():
            return candidate
        if (candidate / "system" / "config.ini").exists():
            return candidate

    if here.parent.name.lower() == "system":
        return here.parent.parent
    return here.parent


def read_config(base_dir: Path) -> configparser.ConfigParser:
    # Only ';' starts an inline comment so that hex colours containing '#' are preserved.
    cfg = configparser.ConfigParser(inline_comment_prefixes=(";",), strict=False)
    for path in (base_dir / "config.ini", base_dir / "system" / "config.ini"):
        if path.exists():
            try:
                cfg.read(path, encoding="utf-8")
            except Exception:
                cfg.read(path)
            break
    if "DEFAULT" not in cfg:
        cfg["DEFAULT"] = {}
    return cfg


def parquet_dir(base_dir: Path, cfg: configparser.ConfigParser) -> Path:
    sub = cfg["DEFAULT"].get("parquet_folder", DEFAULT_PARQUET_SUBDIR)
    folder = (base_dir / sub).resolve()
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def _parquet_search_dirs(base_dir: Path, cfg: configparser.ConfigParser) -> List[Path]:
    sub = cfg["DEFAULT"].get("parquet_folder", DEFAULT_PARQUET_SUBDIR)
    base = base_dir.resolve()
    candidates: List[Path] = []

    def _add(path: Path) -> None:
        try:
            resolved = path.resolve()
        except Exception:
            resolved = path
        if resolved not in candidates:
            candidates.append(resolved)

    _add(base / sub)
    _add(base / "code" / sub)
    _add(base / "code" / "output" / "geoparquet")
    _add(base / "output" / "geoparquet")
    _add(base / "code" / "output")
    _add(base / "output")
    return candidates


def find_parquet_file(base_dir: Path, cfg: configparser.ConfigParser, filename: str) -> Optional[Path]:
    for directory in _parquet_search_dirs(base_dir, cfg):
        candidate = directory / filename
        if candidate.exists():
            return candidate
    return None


def _safe_hex(value: str, fallback: str) -> str:
    text = (value or "").strip()
    if not text:
        text = (fallback or "").strip()
    if text and not text.startswith("#"):
        text = f"#{text}"
    return text or "#BDBDBD"


def read_sensitivity_palette(cfg: configparser.ConfigParser) -> Dict[str, Dict[str, str]]:
    """
    Build a palette mapping sensitivity code -> {"color": str, "description": str}.
    Codes are stored in uppercase, with a dedicated UNKNOWN entry.
    """
    palette: Dict[str, Dict[str, str]] = {}

    default_unknown = _safe_hex(cfg["DEFAULT"].get("category_colour_unknown", ""), "#BDBDBD")
    if cfg.has_section("VALID_VALUES"):
        default_unknown = _safe_hex(cfg["VALID_VALUES"].get("category_colour_unknown", ""), default_unknown)
    if not default_unknown:
        default_unknown = DEFAULT_COLOR_FALLBACK.get(UNKNOWN_CODE, "#8d99ae")

    def _store(code_key: str, color: str, description: str) -> None:
        palette[code_key] = {"color": color, "description": description}

    for code in DEFAULT_SENSITIVITY_ORDER:
        section = cfg[code] if cfg.has_section(code) else None
        color = _safe_hex(section.get("category_colour", "") if section else "", default_unknown)
        description = (section.get("description", "") if section else "").strip()
        _store(code.upper(), color, description)

    _store(UNKNOWN_CODE, default_unknown, "Unknown")
    return palette


# --------------------------------------------------------------------------- #
# Formatting helpers
# --------------------------------------------------------------------------- #

def safe_datetime(value: Any) -> Optional[dt.datetime]:
    if value is None:
        return None
    try:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return None
        if isinstance(ts, pd.Series):
            ts = ts.iloc[0]
        if hasattr(ts, "to_pydatetime"):
            ts = ts.to_pydatetime()
        if isinstance(ts, dt.datetime):
            return ts.replace(tzinfo=None)
    except Exception:
        return None
    return None


def format_timestamp(value: Optional[dt.datetime]) -> str:
    if not value:
        return "--"
    return value.strftime("%Y-%m-%d %H:%M")


def area_to_km2(area_m2: Optional[float]) -> float:
    if area_m2 is None:
        return 0.0
    try:
        value = float(area_m2)
    except Exception:
        return 0.0
    if not math.isfinite(value):
        return 0.0
    if value <= 0:
        return 0.0
    return value / KM2_DENOMINATOR


def format_km2(value_km2: float) -> str:
    return f"{value_km2:,.2f} km^2"


def format_delta(value_km2: float) -> str:
    if not math.isfinite(value_km2) or abs(value_km2) < 0.005:
        return "0.00 km^2"
    sign = "+" if value_km2 > 0 else "-"
    return f"{sign}{abs(value_km2):,.2f} km^2"


def coalesce(*values: Any, default: str = "") -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return default


def _figure_to_png_bytes(fig) -> BytesIO:
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=160, bbox_inches="tight")
    buf.seek(0)
    buf.name = "chart.png"
    return buf


# --------------------------------------------------------------------------- #
# Data access layer
# --------------------------------------------------------------------------- #

@dataclass
class PolygonSummary:
    polygon_id: str
    title: str
    notes: str
    area_m2: float
    has_analysis: bool
    last_timestamp: Optional[dt.datetime]

    @property
    def area_km2(self) -> float:
        return area_to_km2(self.area_m2)

    @property
    def area_label(self) -> str:
        if not self.has_analysis:
            return "-"
        return format_km2(self.area_km2)

    @property
    def timestamp_label(self) -> str:
        return format_timestamp(self.last_timestamp)


@dataclass
class SensitivitySummary:
    code: str
    description: str
    area_m2: float

    @property
    def area_km2(self) -> float:
        return area_to_km2(self.area_m2)

    @property
    def label(self) -> str:
        code_upper = str(self.code).upper()
        display = "Unknown" if code_upper == UNKNOWN_CODE else str(self.code)
        if self.description:
            return f"{display} ({self.description})"
        return display


@dataclass
class AssetGroupSummary:
    asset_group: str
    dominant_code: str
    dominant_description: str
    total_area_m2: float

    @property
    def area_km2(self) -> float:
        return area_to_km2(self.total_area_m2)

    @property
    def area_label(self) -> str:
        return format_km2(self.area_km2)

    @property
    def dominant_label(self) -> str:
        code_display = "Unknown" if self.dominant_code.upper() == UNKNOWN_CODE else self.dominant_code
        if self.dominant_description:
            return f"{code_display} ({self.dominant_description})"
        return code_display


def _polygons_dataframe(polygons: Optional[Iterable[Any]]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for polygon in polygons or []:
        if isinstance(polygon, PolygonSummary):
            source = {
                "Polygon ID": polygon.polygon_id,
                "Title": polygon.title,
                "Notes": polygon.notes,
                "Has analysis": "Yes" if polygon.has_analysis else "No",
                "Area (km²)": polygon.area_km2,
                "Last processed": format_timestamp(polygon.last_timestamp),
            }
        elif isinstance(polygon, dict):
            source = {
                "Polygon ID": polygon.get("polygon_id", ""),
                "Title": polygon.get("title", ""),
                "Notes": polygon.get("notes", ""),
                "Has analysis": "Yes" if polygon.get("has_analysis") else "No",
                "Area (km²)": area_to_km2(polygon.get("area_m2", 0.0)),
                "Last processed": format_timestamp(polygon.get("last_timestamp")),
            }
        else:
            continue
        rows.append(source)
    columns = ["Polygon ID", "Title", "Notes", "Has analysis", "Area (km²)", "Last processed"]
    return pd.DataFrame(rows, columns=columns)


def _sensitivity_dataframe(entries: Optional[Iterable[Any]]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for entry in entries or []:
        if isinstance(entry, SensitivitySummary):
            source = {
                "Code": entry.code,
                "Description": entry.description,
                "Area (km²)": entry.area_km2,
            }
        elif isinstance(entry, dict):
            source = {
                "Code": entry.get("code", ""),
                "Description": entry.get("description", ""),
                "Area (km²)": area_to_km2(entry.get("area_m2", 0.0)),
            }
        else:
            continue
        rows.append(source)
    columns = ["Code", "Description", "Area (km²)"]
    return pd.DataFrame(rows, columns=columns)


def _asset_groups_dataframe(entries: Optional[Iterable[Any]]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for entry in entries or []:
        if isinstance(entry, AssetGroupSummary):
            source = {
                "Asset group": entry.asset_group,
                "Dominant sensitivity": entry.dominant_label,
                "Area (km²)": entry.area_km2,
            }
        elif isinstance(entry, dict):
            code = str(entry.get("dominant_code", "")).upper()
            desc = entry.get("dominant_description", "")
            if code == UNKNOWN_CODE:
                code_display = "Unknown"
            else:
                code_display = code
            dominant = f"{code_display} ({desc})" if desc else code_display
            source = {
                "Asset group": entry.get("asset_group", ""),
                "Dominant sensitivity": dominant,
                "Area (km²)": area_to_km2(entry.get("total_area_m2", 0.0)),
            }
        else:
            continue
        rows.append(source)
    columns = ["Asset group", "Dominant sensitivity", "Area (km²)"]
    return pd.DataFrame(rows, columns=columns)


class AnalysisData:
    """Load and aggregate analysis outputs for comparison."""

    def __init__(self, base_dir: Path, cfg: configparser.ConfigParser, palette_map: Dict[str, Dict[str, str]]) -> None:
        self.base_dir = base_dir
        self.cfg = cfg
        self.palette_map = palette_map or {}
        self.has_sensitivity_max = False
        self.groups = self._load_groups()
        self.polygons = self._load_polygons()
        self.flat = self._load_flat_results()
        self.stacked = self._load_stacked_results()
        self.group_choices = self._build_group_choices()
        self.sensitivity_order = self._build_sensitivity_order()
        self.has_analysis = not self.flat.empty
        self.has_stacked = not self.stacked.empty

    # ----------------------- loading helpers ----------------------- #
    def _load_groups(self) -> pd.DataFrame:
        columns = ["id", "name", "notes", "created_at", "updated_at", "default_geocode"]
        path = find_parquet_file(self.base_dir, self.cfg, ANALYSIS_GROUP_TABLE)
        if not path:
            debug_log(self.base_dir, "analysis groups table not found; continuing with empty frame")
            df = pd.DataFrame(columns=columns)
        else:
            try:
                df = pd.read_parquet(path)
            except Exception as exc:
                debug_log(self.base_dir, f"failed to read {ANALYSIS_GROUP_TABLE}: {exc}")
                df = pd.DataFrame(columns=columns)

        for col in columns:
            if col not in df.columns:
                df[col] = pd.NA

        df["id"] = df["id"].astype(str)
        df["name"] = df["name"].fillna("").astype(str)
        for col in ("created_at", "updated_at"):
            df[col] = pd.to_datetime(df[col], errors="coerce")
        df.sort_values(["name", "id"], inplace=True, ignore_index=True)
        return df

    def _load_polygons(self) -> pd.DataFrame:
        columns = ["id", "group_id", "title", "notes", "created_at", "updated_at"]
        path = find_parquet_file(self.base_dir, self.cfg, ANALYSIS_POLYGON_TABLE)
        if not path:
            debug_log(self.base_dir, "analysis polygons table not found; continuing with empty frame")
            df = pd.DataFrame(columns=columns)
        else:
            try:
                df = gpd.read_parquet(path)
            except Exception as exc:
                debug_log(self.base_dir, f"failed to read {ANALYSIS_POLYGON_TABLE}: {exc}")
                df = pd.DataFrame(columns=columns)

        for col in columns:
            if col not in df.columns:
                df[col] = pd.NA

        df["id"] = df["id"].astype(str)
        df["group_id"] = df["group_id"].astype(str)
        df["title"] = df["title"].fillna("").astype(str)
        df["notes"] = df["notes"].fillna("").astype(str)
        for col in ("created_at", "updated_at"):
            df[col] = pd.to_datetime(df[col], errors="coerce")
        return pd.DataFrame(df)

    def _load_flat_results(self) -> pd.DataFrame:
        path = find_parquet_file(self.base_dir, self.cfg, ANALYSIS_FLAT_TABLE)
        if not path:
            debug_log(self.base_dir, "analysis flat table not found; comparison will show zero results")
            return pd.DataFrame(
                columns=[
                    "analysis_group_id",
                    "analysis_polygon_id",
                    "analysis_polygon_title",
                    "analysis_polygon_notes",
                    "analysis_geocode",
                    "analysis_area_m2",
                    "analysis_timestamp",
                    "sensitivity_code",
                    "sensitivity_description",
                ]
            )
        try:
            gdf = gpd.read_parquet(path)
        except Exception as exc:
            debug_log(self.base_dir, f"failed to read {ANALYSIS_FLAT_TABLE}: {exc}")
            return pd.DataFrame(
                columns=[
                    "analysis_group_id",
                    "analysis_polygon_id",
                    "analysis_polygon_title",
                    "analysis_polygon_notes",
                    "analysis_geocode",
                    "analysis_area_m2",
                    "analysis_timestamp",
                    "sensitivity_code",
                    "sensitivity_description",
                ]
            )

        df = pd.DataFrame(gdf.drop(columns=[col for col in ("geometry",) if col in gdf.columns]))
        self.has_sensitivity_max = "sensitivity_code_max" in gdf.columns
        if "analysis_timestamp" in df.columns:
            df["analysis_timestamp"] = pd.to_datetime(df["analysis_timestamp"], errors="coerce")
        if "analysis_area_m2" in df.columns:
            df["analysis_area_m2"] = pd.to_numeric(df["analysis_area_m2"], errors="coerce")
        for col in ("analysis_group_id", "analysis_polygon_id"):
            if col in df.columns:
                df[col] = df[col].astype(str)
        if "sensitivity_code" in df.columns:
            df["sensitivity_code"] = (
                df["sensitivity_code"]
                .astype(str)
                .str.strip()
                .str.upper()
                .replace({"": "", "NAN": ""})
            )
        else:
            df["sensitivity_code"] = ""
        if "sensitivity_description" in df.columns:
            df["sensitivity_description"] = df["sensitivity_description"].fillna("").astype(str).str.strip()
        else:
            df["sensitivity_description"] = ""

        if "sensitivity_code_max" in df.columns:
            df["sensitivity_code_max"] = (
                df["sensitivity_code_max"]
                .astype(str)
                .str.strip()
                .str.upper()
                .replace({"": "", "NAN": ""})
            )
        else:
            df["sensitivity_code_max"] = ""
        if "sensitivity_description_max" in df.columns:
            df["sensitivity_description_max"] = (
                df["sensitivity_description_max"].fillna("").astype(str).str.strip()
            )
        df["analysis_geocode"] = df.get("analysis_geocode", pd.Series(dtype="object")).fillna("").astype(str)
        df["analysis_polygon_title"] = df.get("analysis_polygon_title", pd.Series(dtype="object")).fillna("").astype(str)
        df["analysis_polygon_notes"] = df.get("analysis_polygon_notes", pd.Series(dtype="object")).fillna("").astype(str)
        return df

    def _load_stacked_results(self) -> pd.DataFrame:
        path = find_parquet_file(self.base_dir, self.cfg, ANALYSIS_STACKED_TABLE)
        if not path:
            debug_log(self.base_dir, "analysis stacked table not found; comprehensive view will start empty")
            return pd.DataFrame(
                columns=[
                    "analysis_group_id",
                    "analysis_polygon_id",
                    "analysis_polygon_title",
                    "analysis_polygon_notes",
                    "analysis_geocode",
                    "analysis_area_m2",
                    "analysis_timestamp",
                    "sensitivity_code",
                    "sensitivity_description",
                    "asset_group_name",
                    "name_gis_assetgroup",
                    "total_asset_objects",
                ]
            )
        try:
            gdf = gpd.read_parquet(path)
        except Exception as exc:
            debug_log(self.base_dir, f"failed to read {ANALYSIS_STACKED_TABLE}: {exc}")
            return pd.DataFrame(
                columns=[
                    "analysis_group_id",
                    "analysis_polygon_id",
                    "analysis_polygon_title",
                    "analysis_polygon_notes",
                    "analysis_geocode",
                    "analysis_area_m2",
                    "analysis_timestamp",
                    "sensitivity_code",
                    "sensitivity_description",
                    "asset_group_name",
                    "name_gis_assetgroup",
                    "total_asset_objects",
                ]
            )
        df = pd.DataFrame(gdf.drop(columns=[col for col in ("geometry",) if col in gdf.columns]))
        if "analysis_timestamp" in df.columns:
            df["analysis_timestamp"] = pd.to_datetime(df["analysis_timestamp"], errors="coerce")
        if "analysis_area_m2" in df.columns:
            df["analysis_area_m2"] = pd.to_numeric(df["analysis_area_m2"], errors="coerce")
        for col in ("analysis_group_id", "analysis_polygon_id"):
            if col in df.columns:
                df[col] = df[col].astype(str)
        if "analysis_polygon_title" in df.columns:
            df["analysis_polygon_title"] = df["analysis_polygon_title"].fillna("").astype(str)
        if "analysis_polygon_notes" in df.columns:
            df["analysis_polygon_notes"] = df["analysis_polygon_notes"].fillna("").astype(str)
        if "analysis_geocode" in df.columns:
            df["analysis_geocode"] = df["analysis_geocode"].fillna("").astype(str)
        if "sensitivity_code" in df.columns:
            df["sensitivity_code"] = (
                df["sensitivity_code"]
                .astype(str)
                .str.strip()
                .str.upper()
                .replace({"": "", "NAN": ""})
            )
        else:
            df["sensitivity_code"] = ""
        if "sensitivity_description" in df.columns:
            df["sensitivity_description"] = df["sensitivity_description"].fillna("").astype(str).str.strip()
        else:
            df["sensitivity_description"] = ""
        for asset_col in ("asset_group_name", "name_gis_assetgroup"):
            if asset_col in df.columns:
                df[asset_col] = df[asset_col].fillna("").astype(str)
        return df

    def _build_group_choices(self) -> List[Tuple[str, str]]:
        choices: List[Tuple[str, str]] = []
        for _, row in self.groups.iterrows():
            group_id = str(row["id"])
            name = row.get("name", "")
            display = name if name else group_id
            display = f"{display} ({group_id})"
            choices.append((group_id, display))
        return choices

    def _polygon_metadata(self, group_id: str) -> Dict[str, Dict[str, Any]]:
        polygon_rows = self.polygons[self.polygons["group_id"].astype(str) == str(group_id)].copy()
        polygon_map: Dict[str, Dict[str, Any]] = {}
        if polygon_rows.empty:
            return polygon_map
        for _, row in polygon_rows.iterrows():
            pid = str(row.get("id", ""))
            if not pid:
                continue
            polygon_map[pid] = {
                "title": coalesce(row.get("title"), default="(untitled)"),
                "notes": row.get("notes", ""),
                "created_at": safe_datetime(row.get("created_at")),
            }
        return polygon_map

    def _build_sensitivity_order(self) -> List[str]:
        order = list(DEFAULT_SENSITIVITY_ORDER)
        seen: set[str] = set(order)
        extras: List[str] = []

        def _add(code: Optional[str]) -> None:
            if not code:
                return
            normalized = str(code).strip().upper()
            if not normalized or normalized in {"NONE", "NAN", UNKNOWN_CODE}:
                return
            if normalized not in seen:
                extras.append(normalized)
                seen.add(normalized)

        for palette_code in self.palette_map.keys():
            _add(palette_code)

        code_column = (
            "sensitivity_code_max"
            if self.has_sensitivity_max and "sensitivity_code_max" in self.flat.columns
            else "sensitivity_code"
        )
        if code_column in self.flat.columns:
            series = (
                self.flat[code_column]
                .dropna()
                .astype(str)
                .str.strip()
            )
            for raw in series:
                _add(raw)

        if extras:
            order.extend(sorted(extras))
        return order

    # ----------------------- aggregation helpers ----------------------- #
    def group_summary(self, group_id: Optional[str]) -> Dict[str, Any]:
        summary = {
            "group_id": group_id or "",
            "group_name": "",
            "group_notes": "",
            "configured_count": 0,
            "processed_count": 0,
            "total_area_m2": 0.0,
            "total_area_label": "0.00 km^2",
            "polygons": [],
            "sensitivity": [],
            "last_run": None,
            "last_run_label": "--",
            "message": "",
        }

        if not group_id:
            summary["message"] = "Select an analysis group."
            return summary

        group_row = self.groups[self.groups["id"].astype(str) == str(group_id)]
        if group_row.empty:
            summary["message"] = "Analysis group not found. Run data_analysis_setup.py first."
            return summary

        summary["group_name"] = coalesce(group_row.iloc[0].get("name"), default=str(group_id))
        summary["group_notes"] = group_row.iloc[0].get("notes", "")

        polygon_map = self._polygon_metadata(str(group_id))

        summary["configured_count"] = len(polygon_map)

        analysis = self.flat[self.flat["analysis_group_id"].astype(str) == str(group_id)].copy()
        if not analysis.empty and "analysis_geocode" in analysis.columns:
            mask = analysis["analysis_geocode"].astype(str).str.lower() == DEFAULT_ANALYSIS_GEOCODE.lower()
            analysis = analysis.loc[mask]

        polygons: List[PolygonSummary] = []
        processed_ids: set[str] = set()
        last_run: Optional[dt.datetime] = None

        if not analysis.empty:
            grouped = analysis.groupby("analysis_polygon_id", dropna=False)
            for polygon_id, subset in grouped:
                pid = str(polygon_id or "").strip()
                if not pid:
                    continue
                processed_ids.add(pid)
                area_m2 = float(subset["analysis_area_m2"].sum(skipna=True))
                ts = subset["analysis_timestamp"].max()
                ts_dt = safe_datetime(ts)
                if ts_dt and (last_run is None or ts_dt > last_run):
                    last_run = ts_dt
                title = coalesce(
                    subset["analysis_polygon_title"].dropna().iloc[-1] if "analysis_polygon_title" in subset else None,
                    polygon_map.get(pid, {}).get("title"),
                    default="(untitled)",
                )
                notes = coalesce(
                    subset["analysis_polygon_notes"].dropna().iloc[-1] if "analysis_polygon_notes" in subset else None,
                    polygon_map.get(pid, {}).get("notes"),
                    default="",
                )
                polygons.append(
                    PolygonSummary(
                        polygon_id=pid,
                        title=title,
                        notes=notes,
                        area_m2=area_m2,
                        has_analysis=True,
                        last_timestamp=ts_dt,
                    )
                )

        for pid, info in polygon_map.items():
            if pid in processed_ids:
                continue
            polygons.append(
                PolygonSummary(
                    polygon_id=pid,
                    title=info.get("title", "(untitled)"),
                    notes=info.get("notes", ""),
                    area_m2=0.0,
                    has_analysis=False,
                    last_timestamp=None,
                )
            )

        polygons.sort(key=lambda p: (0 if p.has_analysis else 1, -p.area_m2))

        summary["polygons"] = polygons
        summary["processed_count"] = len([p for p in polygons if p.has_analysis])
        total_area_m2 = float(sum(p.area_m2 for p in polygons if p.has_analysis))
        summary["total_area_m2"] = total_area_m2
        summary["total_area_label"] = format_km2(area_to_km2(total_area_m2))
        summary["last_run"] = last_run
        summary["last_run_label"] = format_timestamp(last_run)

        totals_by_code: Dict[str, float] = {}
        fallback_desc: Dict[str, str] = {}
        if not analysis.empty:
            code_column = "sensitivity_code_max" if self.has_sensitivity_max and "sensitivity_code_max" in analysis.columns else "sensitivity_code"
            desc_candidates = ["sensitivity_description_max", "sensitivity_description"]
            desc_column = next((col for col in desc_candidates if col in analysis.columns), None)
            if code_column in analysis.columns:
                grouped = analysis.groupby(code_column, dropna=False)
                for code, subset in grouped:
                    raw_code = str(code or "").strip()
                    code_upper = raw_code.upper()
                    if not code_upper or code_upper in {"NONE", "NAN", UNKNOWN_CODE}:
                        continue
                    code_key = code_upper
                    area_m2 = float(subset["analysis_area_m2"].sum(skipna=True))
                    if area_m2 <= 0:
                        continue
                    totals_by_code[code_key] = totals_by_code.get(code_key, 0.0) + area_m2
                    desc = ""
                    if desc_column:
                        desc_values = subset[desc_column].dropna()
                        if not desc_values.empty:
                            desc = str(desc_values.iloc[-1]).strip()
                    if desc and code_key not in fallback_desc:
                        fallback_desc[code_key] = desc
        summary["sensitivity"] = self._sensitivity_entries(totals_by_code, fallback_desc)

        if not totals_by_code and summary["processed_count"] == 0:
            summary["message"] = "No analysis results found. Run the processing workflow in data_analysis_setup.py."
        else:
            summary["message"] = ""
        return summary

    def stacked_summary(self, group_id: Optional[str]) -> Dict[str, Any]:
        summary = {
            "group_id": group_id or "",
            "polygons": [],
            "asset_groups": [],
            "sensitivity": self._sensitivity_entries({}, {}),
            "message": "",
        }

        if not group_id:
            summary["message"] = "Select an analysis group."
            return summary

        if not self.has_stacked:
            summary["message"] = "No comprehensive analysis available yet. Run the processing workflow to populate tbl_analysis_stacked.parquet."
            return summary

        subset = self.stacked[self.stacked["analysis_group_id"].astype(str) == str(group_id)].copy()
        if subset.empty:
            summary["message"] = "No comprehensive analysis results for this group yet."
            return summary

        subset["analysis_area_m2"] = pd.to_numeric(subset["analysis_area_m2"], errors="coerce").fillna(0.0)

        polygon_map = self._polygon_metadata(str(group_id))
        polygon_totals = subset.groupby("analysis_polygon_id", dropna=False)["analysis_area_m2"].sum()
        polygon_times: Dict[str, Optional[dt.datetime]] = {}
        if "analysis_timestamp" in subset.columns:
            ts_series = subset.groupby("analysis_polygon_id", dropna=False)["analysis_timestamp"].max()
            polygon_times = {
                str(pid or "").strip(): safe_datetime(ts_value) for pid, ts_value in ts_series.items()
            }

        polygon_entries: List[PolygonSummary] = []
        for pid_raw, area_m2 in polygon_totals.items():
            pid = str(pid_raw or "").strip()
            if not pid:
                continue
            info = polygon_map.get(pid, {})
            title = info.get("title") or "(untitled)"
            notes = info.get("notes", "")
            last_ts = polygon_times.get(pid)
            polygon_entries.append(
                PolygonSummary(
                    polygon_id=pid,
                    title=title,
                    notes=notes,
                    area_m2=float(area_m2),
                    has_analysis=float(area_m2) > 0.0,
                    last_timestamp=last_ts,
                )
            )

        existing_pids = {p.polygon_id for p in polygon_entries}
        for pid, info in polygon_map.items():
            if pid in existing_pids:
                continue
            polygon_entries.append(
                PolygonSummary(
                    polygon_id=pid,
                    title=info.get("title", "(untitled)"),
                    notes=info.get("notes", ""),
                    area_m2=0.0,
                    has_analysis=False,
                    last_timestamp=None,
                )
            )

        polygon_entries.sort(key=lambda p: (0 if p.has_analysis else 1, -p.area_m2))
        summary["polygons"] = polygon_entries

        totals_by_code: Dict[str, float] = {}
        fallback_desc: Dict[str, str] = {}
        if "sensitivity_code" in subset.columns:
            grouped_codes = subset.groupby("sensitivity_code", dropna=False)
            for code, rows in grouped_codes:
                code_upper = str(code or "").strip().upper()
                if not code_upper or code_upper in {"NONE", "NAN", UNKNOWN_CODE}:
                    continue
                code_key = code_upper
                area_sum = float(rows["analysis_area_m2"].sum(skipna=True))
                if area_sum <= 0:
                    continue
                totals_by_code[code_key] = totals_by_code.get(code_key, 0.0) + area_sum
                if "sensitivity_description" in rows.columns:
                    desc_values = rows["sensitivity_description"].dropna()
                    if not desc_values.empty and code_key not in fallback_desc:
                        fallback_desc[code_key] = str(desc_values.iloc[-1]).strip()

        summary["sensitivity"] = self._sensitivity_entries(totals_by_code, fallback_desc)

        asset_groups: List[AssetGroupSummary] = []
        asset_group_col = None
        for candidate in ("asset_group_name", "name_gis_assetgroup"):
            if candidate in subset.columns:
                asset_group_col = candidate
                break
        if asset_group_col:
            grouped_assets = subset.groupby(asset_group_col, dropna=False)
            for asset_name, rows in grouped_assets:
                total_area = float(rows["analysis_area_m2"].sum(skipna=True))
                if total_area <= 0:
                    continue
                code_totals_series = rows.groupby("sensitivity_code", dropna=False)["analysis_area_m2"].sum()
                valid_totals: Dict[str, float] = {}
                for idx_code, val in code_totals_series.items():
                    code_upper = str(idx_code or "").strip().upper()
                    if not code_upper or code_upper in {"NONE", "NAN", UNKNOWN_CODE}:
                        continue
                    valid_totals[code_upper] = float(val)
                if not valid_totals:
                    continue
                dom_code = max(valid_totals, key=valid_totals.get)
                desc_text = ""
                if "sensitivity_description" in rows.columns:
                    desc_vals = rows["sensitivity_description"].dropna()
                    if not desc_vals.empty:
                        desc_text = str(desc_vals.iloc[-1]).strip()
                palette_desc = self._palette_entry(dom_code).get("description", "")
                dominant_desc = palette_desc or desc_text
                name_display = str(asset_name).strip() or "(unnamed asset group)"
                asset_groups.append(
                    AssetGroupSummary(
                        asset_group=name_display,
                        dominant_code=dom_code,
                        dominant_description=dominant_desc,
                        total_area_m2=total_area,
                    )
                )

        asset_groups.sort(key=lambda item: item.total_area_m2, reverse=True)
        summary["asset_groups"] = asset_groups

        if not totals_by_code:
            summary["message"] = "No overlapping asset results found in the comprehensive analysis."

        return summary

    def _sensitivity_index(self, code: str) -> int:
        code_upper = str(code).upper()
        try:
            return self.sensitivity_order.index(code_upper)
        except ValueError:
            return len(self.sensitivity_order)

    def _palette_entry(self, code: str) -> Dict[str, str]:
        code_upper = str(code).strip().upper() if code is not None else ""
        if not code_upper or code_upper in {"NONE", "NAN"}:
            code_upper = UNKNOWN_CODE
        if code_upper in self.palette_map:
            return self.palette_map[code_upper]
        return self.palette_map.get(UNKNOWN_CODE, {})

    def _sensitivity_entries(
        self,
        totals: Dict[str, float],
        fallback_desc: Dict[str, str],
    ) -> List[SensitivitySummary]:
        entries: List[SensitivitySummary] = []
        for code in self.sensitivity_order:
            palette_entry = self._palette_entry(code)
            description = palette_entry.get("description", "") or fallback_desc.get(code, "")
            area_m2 = float(totals.get(code, 0.0) or 0.0)
            entries.append(SensitivitySummary(code=code, description=description, area_m2=area_m2))
        return entries

    def comparison_rows(
        self,
        left_summary: Optional[Dict[str, Any]],
        right_summary: Optional[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        left_summary = left_summary or {}
        right_summary = right_summary or {}

        def _area(summary: Dict[str, Any]) -> float:
            value = summary.get("total_area_m2", 0.0)
            try:
                return float(value)
            except Exception:
                return 0.0

        rows: List[Dict[str, Any]] = []
        left_total_km2 = area_to_km2(_area(left_summary))
        right_total_km2 = area_to_km2(_area(right_summary))
        rows.append(
            {
                "label": "Total area",
                "left_km2": left_total_km2,
                "right_km2": right_total_km2,
                "delta_km2": left_total_km2 - right_total_km2,
            }
        )

        def _sensitivity_map(summary: Dict[str, Any]) -> Dict[str, SensitivitySummary]:
            mapping: Dict[str, SensitivitySummary] = {}
            for entry in summary.get("sensitivity", []):
                mapping[entry.code] = entry
            return mapping

        left_map = _sensitivity_map(left_summary)
        right_map = _sensitivity_map(right_summary)

        for code in self.sensitivity_order:
            left_entry = left_map.get(code)
            right_entry = right_map.get(code)
            left_km2 = left_entry.area_km2 if left_entry else 0.0
            right_km2 = right_entry.area_km2 if right_entry else 0.0
            description = ""
            if left_entry and left_entry.description:
                description = left_entry.description
            elif right_entry and right_entry.description:
                description = right_entry.description
            palette_desc = self._palette_entry(code).get("description")
            if palette_desc:
                description = palette_desc
            display_code = str(code)
            if description:
                label = f"{display_code} ({description})"
            else:
                label = display_code
            rows.append(
                {
                    "label": label,
                    "left_km2": left_km2,
                    "right_km2": right_km2,
                    "delta_km2": left_km2 - right_km2,
                }
            )
        return rows

    def status_message(self) -> str:
        if self.groups.empty:
            return "No analysis groups available. Run data_analysis_setup.py to create them."
        if not self.has_analysis:
            return (
                "No analysis results detected. Run the clipping workflow in data_analysis_setup.py "
                "to populate tbl_analysis_flat.parquet."
            )
        return f"Loaded {len(self.groups)} group(s). Select two groups to compare."


# --------------------------------------------------------------------------- #
# UI components
# --------------------------------------------------------------------------- #


class GroupHeader:
    """Selector and headline statistics for a comparison column."""

    def __init__(self, master: tk.Widget, title: str, on_selection: callable) -> None:
        self._on_selection = on_selection
        self.frame = ttk.Frame(master, padding=(6, 6))
        self.frame.columnconfigure(0, weight=1)

        self.header_label = ttk.Label(self.frame, text=title, font=("Segoe UI", 12, "bold"))
        apply_bootstyle(self.header_label, PRIMARY)
        self.header_label.grid(row=0, column=0, sticky="w")

        self.combo_var = tk.StringVar()
        self.combobox = ttk.Combobox(self.frame, textvariable=self.combo_var, state="readonly", width=40)
        self.combobox.grid(row=1, column=0, sticky="ew", pady=(2, 4))
        self.combobox.bind("<<ComboboxSelected>>", self._handle_selection)
        apply_bootstyle(self.combobox, SECONDARY)

        summary_frame = ttk.Frame(self.frame)
        summary_frame.grid(row=2, column=0, sticky="ew")
        summary_frame.columnconfigure(1, weight=1)

        ttk.Label(summary_frame, text="Total area:", style="Caption.TLabel").grid(row=0, column=0, sticky="w")
        self.total_var = tk.StringVar(value="0.00 km^2")
        ttk.Label(summary_frame, textvariable=self.total_var, style="Value.TLabel").grid(row=0, column=1, sticky="w")

        ttk.Label(summary_frame, text="Polygons processed:", style="Caption.TLabel").grid(row=1, column=0, sticky="w")
        self.count_var = tk.StringVar(value="0 / 0")
        ttk.Label(summary_frame, textvariable=self.count_var, style="Value.TLabel").grid(row=1, column=1, sticky="w")

        ttk.Label(summary_frame, text="Last run:", style="Caption.TLabel").grid(row=2, column=0, sticky="w")
        self.last_run_var = tk.StringVar(value="--")
        ttk.Label(summary_frame, textvariable=self.last_run_var, style="Value.TLabel").grid(row=2, column=1, sticky="w")

    def _handle_selection(self, _event: object = None) -> None:
        if self._on_selection:
            self._on_selection(self.combo_var.get())

    def set_options(self, options: List[str]) -> None:
        self.combobox["values"] = options
        if not options:
            self.combo_var.set("")

    def set_selection(self, display: Optional[str]) -> None:
        self.combo_var.set(display or "")

    def current_selection(self) -> str:
        return self.combo_var.get()

    def update_summary(self, summary: Dict[str, Any]) -> None:
        summary = summary or {}
        self.total_var.set(summary.get("total_area_label", "0.00 km^2"))
        processed = summary.get("processed_count", 0)
        configured = summary.get("configured_count", 0)
        self.count_var.set(f"{processed} / {configured}")
        self.last_run_var.set(summary.get("last_run_label", "--"))


class SensitivityChart:
    """Reusable horizontal bar chart for sensitivity data."""

    def __init__(self, master: tk.Widget, palette_map: Dict[str, Dict[str, str]]) -> None:
        self.palette_map = palette_map or {}
        master.columnconfigure(0, weight=1)
        master.rowconfigure(0, weight=1)
        # Fixed footprint ~340x260 px to balance size and readability
        self.figure = Figure(figsize=(3.4, 2.6), dpi=100)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_xlabel("Area (km^2)", fontsize=9)
        self.ax.set_ylabel("Sensitivity", fontsize=9)
        self.ax.tick_params(axis="both", labelsize=9)
        self.ax.grid(True, axis="x", linestyle=":", linewidth=0.5, alpha=0.6)

        holder = ttk.Frame(master)
        holder.grid(row=0, column=0, sticky="n")
        holder.columnconfigure(0, weight=1)
        holder.columnconfigure(1, weight=0)
        holder.columnconfigure(2, weight=1)

        self.canvas = FigureCanvasTkAgg(self.figure, master=holder)
        self.canvas.get_tk_widget().grid(row=0, column=1, sticky="n")

    def _palette_entry(self, code: str) -> Dict[str, str]:
        code_upper = str(code).strip().upper() if code is not None else ""
        if not code_upper or code_upper in {"NONE", "NAN"}:
            code_upper = UNKNOWN_CODE
        return self.palette_map.get(code_upper, {})

    def _color_for(self, code: str) -> str:
        entry = self._palette_entry(code)
        if entry and entry.get("color"):
            return entry["color"]
        return DEFAULT_COLOR_FALLBACK.get(str(code).strip().upper(), "#4e79a7")

    def update(self, data: Iterable[SensitivitySummary]) -> None:
        self.ax.clear()
        self.ax.set_xlabel("Area (km^2)", fontsize=9)
        self.ax.set_ylabel("Sensitivity", fontsize=9)
        self.ax.tick_params(axis="both", labelsize=9)
        self.ax.grid(True, axis="x", linestyle=":", linewidth=0.5, alpha=0.6)

        entries = list(data)
        if not entries:
            self.ax.text(
                0.5,
                0.5,
                "No sensitivity data",
                ha="center",
                va="center",
                fontsize=10,
                transform=self.ax.transAxes,
            )
            self.figure.tight_layout()
            self.canvas.draw_idle()
            return

        values = [entry.area_km2 for entry in entries]
        labels = [entry.label for entry in entries]
        y_positions = np.arange(len(entries))
        colors = [self._color_for(entry.code) for entry in entries]

        self.ax.barh(y_positions, values, color=colors)
        self.ax.set_yticks(y_positions)
        self.ax.set_yticklabels(labels)
        self.ax.set_xlim(left=0)
        self.ax.invert_yaxis()

        max_value = max(values) if values else 0.0
        offset = max(0.03 * max_value, 0.01)
        for idx, value in enumerate(values):
            self.ax.text(
                value + offset,
                y_positions[idx],
                f"{value:,.2f}",
                va="center",
                fontsize=8,
            )

        self.figure.tight_layout()
        self.canvas.draw_idle()


class PolygonTotalsPanel:
    """Displays polygon totals from stacked analysis."""

    def __init__(self, master: tk.Widget) -> None:
        self.frame = ttk.Frame(master, padding=(6, 6))
        self.frame.columnconfigure(0, weight=1)
        # Avoid stretching the chart; keep minimal height
        self.frame.rowconfigure(1, weight=0)

        self.message_var = tk.StringVar(value="")
        self.message_label = ttk.Label(self.frame, textvariable=self.message_var, wraplength=360, justify="left")
        apply_bootstyle(self.message_label, WARNING)
        self.message_label.grid(row=0, column=0, sticky="w")
        self.message_label.grid_remove()

        polygon_frame = ttk.LabelFrame(self.frame, text="Polygon totals (stacked)")
        polygon_frame.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        polygon_frame.columnconfigure(0, weight=1)
        polygon_frame.rowconfigure(0, weight=1)
        apply_bootstyle(polygon_frame, SECONDARY)

        self.tree = ttk.Treeview(
            polygon_frame,
            columns=("title", "area"),
            show="headings",
            height=8,
        )
        self.tree.heading("title", text="Title")
        self.tree.heading("area", text="Area (km^2)")
        self.tree.column("title", width=240, anchor="w")
        self.tree.column("area", width=120, anchor="e")
        self.tree.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(polygon_frame, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

    def update(self, summary: Dict[str, Any]) -> None:
        summary = summary or {}
        message = summary.get("message", "")
        self.message_var.set(message)
        if message:
            self.message_label.grid()
        else:
            self.message_label.grid_remove()

        for item in self.tree.get_children():
            self.tree.delete(item)
        for polygon in summary.get("polygons", []):
            self.tree.insert(
                "",
                "end",
                values=(polygon.title, polygon.area_label),
            )


class AssetOverlapPanel:
    """Displays asset group overlap from stacked analysis."""

    def __init__(self, master: tk.Widget) -> None:
        self.frame = ttk.Frame(master, padding=(6, 6))
        self.frame.columnconfigure(0, weight=1)
        self.frame.rowconfigure(1, weight=1)

        self.message_var = tk.StringVar(value="")
        self.message_label = ttk.Label(self.frame, textvariable=self.message_var, wraplength=360, justify="left")
        apply_bootstyle(self.message_label, WARNING)
        self.message_label.grid(row=0, column=0, sticky="w")
        self.message_label.grid_remove()

        asset_frame = ttk.LabelFrame(self.frame, text="Asset group overlap")
        asset_frame.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        asset_frame.columnconfigure(0, weight=1)
        asset_frame.rowconfigure(0, weight=1)
        apply_bootstyle(asset_frame, SUCCESS)

        self.tree = ttk.Treeview(
            asset_frame,
            columns=("asset", "sensitivity", "area"),
            show="headings",
            height=10,
        )
        self.tree.heading("asset", text="Asset group")
        self.tree.heading("sensitivity", text="Dominant sensitivity")
        self.tree.heading("area", text="Area (km^2)")
        self.tree.column("asset", width=240, anchor="w")
        self.tree.column("sensitivity", width=200, anchor="w")
        self.tree.column("area", width=120, anchor="e")
        self.tree.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(asset_frame, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

    def update(self, summary: Dict[str, Any]) -> None:
        summary = summary or {}
        message = summary.get("message", "")
        self.message_var.set(message)
        if message:
            self.message_label.grid()
        else:
            self.message_label.grid_remove()

        for item in self.tree.get_children():
            self.tree.delete(item)
        for asset in summary.get("asset_groups", []):
            self.tree.insert(
                "",
                "end",
                values=(asset.asset_group, asset.dominant_label, asset.area_label),
            )


class SensitivityTotalsPanel:
    """Displays sensitivity totals chart from stacked analysis."""

    def __init__(self, master: tk.Widget, palette_map: Dict[str, Dict[str, str]]) -> None:
        self.frame = ttk.Frame(master, padding=(6, 6))
        self.frame.columnconfigure(0, weight=1)
        self.frame.rowconfigure(1, weight=1)

        self.message_var = tk.StringVar(value="")
        self.message_label = ttk.Label(self.frame, textvariable=self.message_var, wraplength=360, justify="left")
        apply_bootstyle(self.message_label, WARNING)
        self.message_label.grid(row=0, column=0, sticky="w")
        self.message_label.grid_remove()

        chart_frame = ttk.LabelFrame(self.frame, text="Sensitivity totals")
        chart_frame.grid(row=1, column=0, sticky="nw", pady=(4, 0))
        apply_bootstyle(chart_frame, INFO)
        chart_frame.columnconfigure(0, weight=0)
        chart_frame.rowconfigure(0, weight=0)
        self.chart = SensitivityChart(chart_frame, palette_map)

    def update(self, summary: Dict[str, Any]) -> None:
        summary = summary or {}
        message = summary.get("message", "")
        self.message_var.set(message)
        if message:
            self.message_label.grid()
        else:
            self.message_label.grid_remove()
        self.chart.update(summary.get("sensitivity", []))

    def chart_figure(self):
        return self.chart.figure


class TwoColumnView:
    """Generic left/right wrapper to keep side-by-side layout."""

    def __init__(self, master: tk.Widget, left_factory, right_factory) -> None:
        self.frame = ttk.Frame(master)
        self.frame.columnconfigure(0, weight=1)
        self.frame.columnconfigure(1, weight=1)
        self.frame.rowconfigure(0, weight=1)

        left_holder = ttk.Frame(self.frame)
        left_holder.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        left_holder.columnconfigure(0, weight=1)
        left_holder.rowconfigure(0, weight=1)

        right_holder = ttk.Frame(self.frame)
        right_holder.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        right_holder.columnconfigure(0, weight=1)
        right_holder.rowconfigure(0, weight=1)

        self.left = left_factory(left_holder)
        self.right = right_factory(right_holder)
        # Ensure child frames are placed inside their holders
        if hasattr(self.left, "frame"):
            self.left.frame.grid(row=0, column=0, sticky="nsew")
        if hasattr(self.right, "frame"):
            self.right.frame.grid(row=0, column=0, sticky="nsew")


class PolygonTotalsView:
    def __init__(self, master: tk.Widget) -> None:
        columns = TwoColumnView(
            master,
            lambda parent: PolygonTotalsPanel(parent),
            lambda parent: PolygonTotalsPanel(parent),
        )
        self.frame = columns.frame
        self.left = columns.left
        self.right = columns.right

    def update_left(self, summary: Dict[str, Any]) -> None:
        self.left.update(summary)

    def update_right(self, summary: Dict[str, Any]) -> None:
        self.right.update(summary)


class AssetOverlapView:
    def __init__(self, master: tk.Widget) -> None:
        columns = TwoColumnView(
            master,
            lambda parent: AssetOverlapPanel(parent),
            lambda parent: AssetOverlapPanel(parent),
        )
        self.frame = columns.frame
        self.left = columns.left
        self.right = columns.right

    def update_left(self, summary: Dict[str, Any]) -> None:
        self.left.update(summary)

    def update_right(self, summary: Dict[str, Any]) -> None:
        self.right.update(summary)


class SensitivityTotalsView:
    def __init__(self, master: tk.Widget, palette_map: Dict[str, Dict[str, str]]) -> None:
        columns = TwoColumnView(
            master,
            lambda parent: SensitivityTotalsPanel(parent, palette_map),
            lambda parent: SensitivityTotalsPanel(parent, palette_map),
        )
        self.frame = columns.frame
        self.left = columns.left
        self.right = columns.right

    def update_left(self, summary: Dict[str, Any]) -> None:
        self.left.update(summary)

    def update_right(self, summary: Dict[str, Any]) -> None:
        self.right.update(summary)

    def chart_figures(self) -> List[Tuple[str, Any]]:
        return [
            ("Sensitivity totals - Left", self.left.chart_figure()),
            ("Sensitivity totals - Right", self.right.chart_figure()),
        ]


class ComparisonTotalsView:
    def __init__(self, master: tk.Widget) -> None:
        self.frame = ttk.Frame(master, padding=(6, 6))
        self.frame.columnconfigure(0, weight=1)
        self.frame.rowconfigure(0, weight=1)
        self.table = ComparisonTable(self.frame)
        self.table.frame.grid(row=0, column=0, sticky="nsew")

    def update_rows(self, rows: Iterable[Dict[str, Any]]) -> None:
        self.table.update_rows(rows)


class ComparisonTable:
    def __init__(self, master: tk.Widget) -> None:
        self.frame = ttk.LabelFrame(master, text="Comparison totals")
        self.frame.columnconfigure(0, weight=1)
        self.frame.rowconfigure(0, weight=1)
        apply_bootstyle(self.frame, INFO)

        self.tree = ttk.Treeview(
            self.frame,
            columns=("left", "right", "delta"),
            show="tree headings",
            height=6,
        )
        self.tree.heading("#0", text="Category")
        self.tree.heading("left", text="Left (km^2)")
        self.tree.heading("right", text="Right (km^2)")
        self.tree.heading("delta", text="Difference (L - R)")
        self.tree.column("#0", width=300, anchor="w")
        self.tree.column("left", width=120, anchor="e")
        self.tree.column("right", width=120, anchor="e")
        self.tree.column("delta", width=140, anchor="e")
        self.tree.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(self.frame, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

    def update_rows(self, rows: Iterable[Dict[str, Any]]) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)
        row_list = list(rows)
        self.tree.config(height=max(6, len(row_list) or 1))
        for row in row_list:
            self.tree.insert(
                "",
                "end",
                text=row.get("label", ""),
                values=(
                    format_km2(row.get("left_km2", 0.0)),
                    format_km2(row.get("right_km2", 0.0)),
                    format_delta(row.get("delta_km2", 0.0)),
                ),
            )


class ComparisonApp:
    def __init__(self, data: AnalysisData, theme: str) -> None:
        self.data = data
        self.palette_map = data.palette_map
        self.root = tb.Window(themename=theme)
        self.root.title("MESA Area Analysis - Comparison")
        self.root.geometry("1280x780")
        self.root.minsize(1100, 720)

        self._configure_style()

        container = ttk.Frame(self.root, padding=(12, 12))
        container.pack(fill=tk.BOTH, expand=True)
        apply_bootstyle(container, SECONDARY)

        header_frame = ttk.Frame(container)
        header_frame.pack(fill=tk.X, pady=(0, 6))
        header_frame.columnconfigure(0, weight=1)
        header_frame.columnconfigure(1, weight=1)
        apply_bootstyle(header_frame, SECONDARY)

        self.left_header = GroupHeader(header_frame, "Left group", self._on_left_selection)
        self.left_header.frame.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        self.right_header = GroupHeader(header_frame, "Right group", self._on_right_selection)
        self.right_header.frame.grid(row=0, column=1, sticky="ew", padx=(8, 0))

        self.notebook = ttk.Notebook(container)
        apply_bootstyle(self.notebook, SECONDARY)
        self.polygon_view = PolygonTotalsView(self.notebook)
        self.notebook.add(self.polygon_view.frame, text="Polygon totals (stacked)")
        self.asset_view = AssetOverlapView(self.notebook)
        self.notebook.add(self.asset_view.frame, text="Asset group overlap")
        self.sensitivity_view = SensitivityTotalsView(self.notebook, self.palette_map)
        self.notebook.add(self.sensitivity_view.frame, text="Sensitivity totals")
        self.comparison_view = ComparisonTotalsView(self.notebook)
        self.notebook.add(self.comparison_view.frame, text="Comparison totals")
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=(6, 0))

        self.status_var = tk.StringVar(value=self.data.status_message())
        self.status_label = ttk.Label(container, textvariable=self.status_var, anchor="w")
        apply_bootstyle(self.status_label, INFO)
        self.status_label.pack(fill=tk.X, pady=(8, 0))

        self.report_dir = (self.data.base_dir / "output" / "reports").resolve()
        self.report_note_var = tk.StringVar(value="")
        self.report_link_var = tk.StringVar(value="")
        self.report_note_label = ttk.Label(container, textvariable=self.report_note_var, anchor="w")
        apply_bootstyle(self.report_note_label, SECONDARY)
        self.report_link = tk.Label(
            container,
            textvariable=self.report_link_var,
            fg="#0d6efd",
            cursor="hand2",
            anchor="w",
        )
        self.report_link.bind("<Button-1>", self._open_report_folder)

        footer = ttk.Frame(container)
        footer.pack(fill=tk.X, pady=(12, 0))
        footer.columnconfigure(0, weight=1)
        footer.columnconfigure(1, weight=1)
        tb.Button(footer, text="Print report", bootstyle=SUCCESS, command=self._export_comprehensive_report).grid(
            row=0, column=0, sticky="ew", padx=(0, 6), pady=(0, 6)
        )
        tb.Button(footer, text="Exit", bootstyle=PRIMARY, command=self.root.destroy).grid(
            row=0, column=1, sticky="ew", padx=(6, 0), pady=(0, 6)
        )

        self.display_to_id: Dict[str, str] = {}
        self.left_summary: Dict[str, Any] = {}
        self.right_summary: Dict[str, Any] = {}
        self.left_stacked_summary: Dict[str, Any] = {}
        self.right_stacked_summary: Dict[str, Any] = {}

        self._populate_options()

    def _configure_style(self) -> None:
        style = ttk.Style(self.root)
        style.configure("Caption.TLabel", font=("Segoe UI", 9))
        style.configure("Value.TLabel", font=("Segoe UI", 10, "bold"))

    def _populate_options(self) -> None:
        choices = self.data.group_choices
        option_labels = [display for _, display in choices]
        self.display_to_id = {display: group_id for group_id, display in choices}

        self.left_header.set_options(option_labels)
        self.right_header.set_options(option_labels)

        empty_basic = self.data.group_summary(None)
        empty_comp = self.data.stacked_summary(None)

        if not option_labels:
            self.left_header.update_summary(empty_basic)
            self.right_header.update_summary(empty_basic)
            self.polygon_view.update_left(empty_comp)
            self.polygon_view.update_right(empty_comp)
            self.asset_view.update_left(empty_comp)
            self.asset_view.update_right(empty_comp)
            self.sensitivity_view.update_left(empty_comp)
            self.sensitivity_view.update_right(empty_comp)
            self.comparison_view.update_rows([])
            self.status_var.set(self.data.status_message())
            self.left_summary = empty_basic
            self.right_summary = empty_basic
            self.left_stacked_summary = empty_comp
            self.right_stacked_summary = empty_comp
            return

        left_display = option_labels[0]
        self.left_header.set_selection(left_display)
        self._on_left_selection(left_display)

        right_display = option_labels[1] if len(option_labels) > 1 else option_labels[0]
        self.right_header.set_selection(right_display)
        self._on_right_selection(right_display)

        if self.data.has_analysis:
            self.status_var.set("Ready. Adjust the selections to compare different groups.")
        else:
            self.status_var.set(self.data.status_message())

    def _on_left_selection(self, display: str) -> None:
        group_id = self.display_to_id.get(display)
        self.left_summary = self.data.group_summary(group_id)
        self.left_header.update_summary(self.left_summary)
        self.left_stacked_summary = self.data.stacked_summary(group_id)
        self.polygon_view.update_left(self.left_stacked_summary)
        self.asset_view.update_left(self.left_stacked_summary)
        self.sensitivity_view.update_left(self.left_stacked_summary)
        self._refresh_comparison()

    def _on_right_selection(self, display: str) -> None:
        group_id = self.display_to_id.get(display)
        self.right_summary = self.data.group_summary(group_id)
        self.right_header.update_summary(self.right_summary)
        self.right_stacked_summary = self.data.stacked_summary(group_id)
        self.polygon_view.update_right(self.right_stacked_summary)
        self.asset_view.update_right(self.right_stacked_summary)
        self.sensitivity_view.update_right(self.right_stacked_summary)
        self._refresh_comparison()

    def _refresh_comparison(self) -> None:
        rows = self.data.comparison_rows(self.left_summary, self.right_summary)
        self.comparison_view.update_rows(rows)

        messages: List[str] = []
        for summary in (
            self.left_summary,
            self.right_summary,
            self.left_stacked_summary,
            self.right_stacked_summary,
        ):
            if isinstance(summary, dict):
                msg = summary.get("message")
                if msg:
                    messages.append(str(msg))
        if messages:
            self.status_var.set(messages[0])
        else:
            self.status_var.set(self.data.status_message())

    def _ensure_report_dir(self) -> Path:
        self.report_dir.mkdir(parents=True, exist_ok=True)
        return self.report_dir

    def _basic_overview_df(self) -> pd.DataFrame:
        left = self.left_summary or {}
        right = self.right_summary or {}
        rows = [
            {"Metric": "Group name", "Left": left.get("group_name", ""), "Right": right.get("group_name", "")},
            {"Metric": "Notes", "Left": left.get("group_notes", ""), "Right": right.get("group_notes", "")},
            {"Metric": "Configured polygons", "Left": left.get("configured_count", 0), "Right": right.get("configured_count", 0)},
            {"Metric": "Processed polygons", "Left": left.get("processed_count", 0), "Right": right.get("processed_count", 0)},
            {"Metric": "Total analysed area (km²)", "Left": area_to_km2(left.get("total_area_m2", 0.0)), "Right": area_to_km2(right.get("total_area_m2", 0.0))},
            {"Metric": "Last run", "Left": left.get("last_run_label", "--"), "Right": right.get("last_run_label", "--")},
        ]
        return pd.DataFrame(rows, columns=["Metric", "Left", "Right"])

    def _comprehensive_overview_df(self) -> pd.DataFrame:
        def _sum_polygons(summary: Dict[str, Any]) -> float:
            total = 0.0
            for polygon in summary.get("polygons", []) or []:
                if isinstance(polygon, PolygonSummary):
                    total += polygon.area_km2
                elif isinstance(polygon, dict):
                    total += area_to_km2(polygon.get("area_m2", 0.0))
            return total

        def _asset_count(summary: Dict[str, Any]) -> int:
            assets = summary.get("asset_groups") or []
            return len(list(assets))

        left = self.left_stacked_summary or {}
        right = self.right_stacked_summary or {}
        rows = [
            {"Metric": "Group name", "Left": left.get("group_id", self.left_summary.get("group_name", "")), "Right": right.get("group_id", self.right_summary.get("group_name", ""))},
            {"Metric": "Polygons with overlaps", "Left": len(left.get("polygons", []) or []), "Right": len(right.get("polygons", []) or [])},
            {"Metric": "Total overlap area (km²)", "Left": _sum_polygons(left), "Right": _sum_polygons(right)},
            {"Metric": "Asset groups reported", "Left": _asset_count(left), "Right": _asset_count(right)},
        ]
        return pd.DataFrame(rows, columns=["Metric", "Left", "Right"])

    def _comparison_dataframe(self) -> pd.DataFrame:
        rows = self.data.comparison_rows(self.left_summary, self.right_summary)
        data_rows: List[Dict[str, Any]] = []
        for row in rows:
            data_rows.append(
                {
                    "Category": row.get("label", ""),
                    "Left (km²)": row.get("left_km2", 0.0),
                    "Right (km²)": row.get("right_km2", 0.0),
                    "Difference (km²)": row.get("delta_km2", 0.0),
                }
            )
        columns = ["Category", "Left (km²)", "Right (km²)", "Difference (km²)"]
        return pd.DataFrame(data_rows, columns=columns)

    def _write_excel_report(
        self,
        sheets: Dict[str, pd.DataFrame],
        prefix: str,
        chart_figures: Optional[List[Tuple[str, Any]]] = None,
    ) -> Path:
        report_dir = self._ensure_report_dir()
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = report_dir / f"{prefix}_{timestamp}.xlsx"
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, frame in sheets.items():
                safe_name = sheet_name[:31] or "Sheet"
                df = frame if not frame.empty else pd.DataFrame(columns=frame.columns if len(frame.columns) else ["Data"])
                df.to_excel(writer, sheet_name=safe_name, index=False)
            if chart_figures:
                charts_sheet = writer.book.create_sheet("Charts")
                row_cursor = 1
                for title, fig in chart_figures:
                    charts_sheet.cell(row=row_cursor, column=1, value=title)
                    row_cursor += 1
                    try:
                        stream = _figure_to_png_bytes(fig)
                        stream.name = "chart.png"
                        img = XLImage(stream)
                        img.anchor = f"A{row_cursor}"
                        charts_sheet.add_image(img)
                        height_rows = max(int((img.height or 240) / 18), 15)
                        row_cursor += height_rows + 2
                    except Exception as exc:
                        charts_sheet.cell(row=row_cursor, column=1, value=f"(chart unavailable: {exc})")
                        row_cursor += 4
        return path

    def _after_report_export(self, path: Path) -> None:
        msg = f"Report exported: {path.name}"
        self.report_note_var.set(msg)
        self.report_link_var.set(f"Open reports folder ({self.report_dir})")
        if not self.report_note_label.winfo_ismapped():
            self.report_note_label.pack(fill=tk.X, pady=(4, 0))
        if not self.report_link.winfo_ismapped():
            self.report_link.pack(anchor="w", pady=(0, 8))
        debug_log(self.data.base_dir, msg)

    def _open_report_folder(self, _event=None) -> None:
        try:
            self._ensure_report_dir()
            folder = str(self.report_dir)
            if sys.platform.startswith("win"):
                os.startfile(folder)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        except Exception as exc:
            messagebox.showerror("Open folder failed", str(exc))

    def _ensure_selection(self, summary: Dict[str, Any]) -> bool:
        return bool(summary and summary.get("group_id"))

    def _export_basic_report(self) -> None:
        if not (self._ensure_selection(self.left_summary) and self._ensure_selection(self.right_summary)):
            messagebox.showwarning("Export basic report", "Select two analysis groups first.")
            return
        try:
            sheets = {
                "Overview": self._basic_overview_df(),
                "Left polygons": _polygons_dataframe(self.left_summary.get("polygons")),
                "Right polygons": _polygons_dataframe(self.right_summary.get("polygons")),
                "Left sensitivity": _sensitivity_dataframe(self.left_summary.get("sensitivity")),
                "Right sensitivity": _sensitivity_dataframe(self.right_summary.get("sensitivity")),
                "Comparison": self._comparison_dataframe(),
            }
            charts = self.sensitivity_view.chart_figures()
            path = self._write_excel_report(sheets, "basic_analysis", chart_figures=charts)
            self._after_report_export(path)
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))
            debug_log(self.data.base_dir, f"basic report export failed: {exc}")

    def _export_comprehensive_report(self) -> None:
        if not (self._ensure_selection(self.left_stacked_summary) and self._ensure_selection(self.right_stacked_summary)):
            messagebox.showwarning("Export comprehensive report", "Select groups that have comprehensive analysis results first.")
            return
        try:
            sheets = {
                "Overview": self._comprehensive_overview_df(),
                "Left polygons": _polygons_dataframe(self.left_stacked_summary.get("polygons")),
                "Right polygons": _polygons_dataframe(self.right_stacked_summary.get("polygons")),
                "Left assets": _asset_groups_dataframe(self.left_stacked_summary.get("asset_groups")),
                "Right assets": _asset_groups_dataframe(self.right_stacked_summary.get("asset_groups")),
                "Left sensitivity": _sensitivity_dataframe(self.left_stacked_summary.get("sensitivity")),
                "Right sensitivity": _sensitivity_dataframe(self.right_stacked_summary.get("sensitivity")),
            }
            charts = self.sensitivity_view.chart_figures()
            path = self._write_excel_report(sheets, "comprehensive_analysis", chart_figures=charts)
            self._after_report_export(path)
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))
            debug_log(self.data.base_dir, f"comprehensive report export failed: {exc}")

    def run(self) -> None:
        self.root.mainloop()


# --------------------------------------------------------------------------- #
# CLI entry point
# --------------------------------------------------------------------------- #


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare analysis groups side by side.")
    parser.add_argument(
        "--original_working_directory",
        dest="owd",
        help="Path to the working directory (Mesa desktop supplies this automatically).",
    )
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> None:
    args = parse_args(argv)
    base_dir = resolve_base_dir(args.owd)
    cfg = read_config(base_dir)

    palette_map = read_sensitivity_palette(cfg)
    for code_key, entry in palette_map.items():
        debug_log(base_dir, f"presentation palette {code_key}: {entry.get('color', '')}")

    theme = cfg["DEFAULT"].get("ttk_bootstrap_theme", "litera").strip() or "litera"
    debug_log(base_dir, f"presentation theme: {theme}")

    data = AnalysisData(base_dir, cfg, palette_map)
    debug_log(base_dir, "Comparison viewer initialised")

    app = ComparisonApp(data, theme)
    app.run()


if __name__ == "__main__":
    main()
